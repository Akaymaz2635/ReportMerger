import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_csv_files():

    # İşlem yapılacak klasör yolu
    input_folder = "C:/Users/alika/Desktop/Reporter_Project/Inputs"

    # Tüm .csv dosyalarını al
    csv_files = glob.glob(input_folder + "/*.csv")

    # İlk .csv dosyasını oku ve 'Min' ve 'Max' sütunlarını ekle
    merged_df = pd.read_csv(csv_files[0])
    merged_df.insert(loc=7, column='Min', value='')
    merged_df.insert(loc=8, column='Max', value='')

    # 'Dev' ve 'Out' sütunlarını sil
    merged_df = merged_df.drop(columns=['Dev' , 'Actual' , 'Out' , 'Check'])

    # Diğer .csv dosyalarını oku ve 'Check' sütunlarını ekleyin
    for file in csv_files[0:]:
        # Dosya adını al ve 'Check' sütununu oku
        column_name = file.split("/")[-1].split(".csv")[0][-1]  # Dosya adının son karakteri
        check_column = pd.read_csv(file)['Check']
        # DataFrame'e 'Check' sütununu ekle
        merged_df[column_name] = check_column

    # 'Min' ve 'Max' sütunlarındaki boş hücrelere minimum ve maksimum değerleri yaz
    for index, row in merged_df.iterrows():
        min_value = row[9:].min()  # İlk 10 sütunun dışındaki sütunlar 'Check' sütunlarıdır
        max_value = row[9:].max()
        merged_df.at[index, 'Min'] = min_value
        merged_df.at[index, 'Max'] = max_value

    # Kolon adlarını güncelle
    columns_mapping = {col: col.split(" ")[0] for col in merged_df.columns[9:]}
    merged_df = merged_df.rename(columns=columns_mapping)

    # Birleştirilmiş veriyi göster
    print(merged_df)

    # Birleştirilmiş veriyi Excel dosyasına dönüştür ve kaydet
    output_path = "C:/Users/alika/Desktop/Reporter_Project/Outputs/merged_data.xlsx"
    merged_df.to_excel(output_path, index=False, engine='openpyxl', float_format="%.2f", header=True)

    # Excel dosyasını aç ve hücre tiplerini belirle
    wb = load_workbook(output_path)
    ws = wb.active

    # Mavi ve kırmızı renkler
    blue_fill = PatternFill(start_color="00CCFF", end_color="00CCFF", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Her satır için koşulları kontrol et ve hücreleri uygun şekilde boyayın
    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
        for index, cell in enumerate(row[2:], start=2):  # 'Min' kolonundan başlayarak
            nominal = float(row[2].value)  # Nominal değeri al
            tol_minus = float(row[3].value)  # Tol- değeri al
            tol_plus = float(row[4].value)  # Tol+ değeri al
            cell_value = row[index].value  # Hücrenin değeri
            try: #çalışmıyor
                cell_value = float(cell_value)
                if cell_value < (nominal - tol_minus):  # Koşul 1: Nominal - Tol-
                    cell.fill = blue_fill  # Maviye boyama
                elif cell_value > (nominal + tol_plus):  # Koşul 2: Nominal + Tol+
                    cell.fill = red_fill  # Kırmızıya boyama
            except ValueError:
                pass

    # Excel dosyasını kaydet
    wb.save(output_path)

process_csv_files()

def apply_inverse_transformation(input_excel, output_excel):
    # Excel dosyasını oku
    df = pd.read_excel(input_excel)

    # İstenmeyen sütunları sil
    df = df.drop(columns=['Property', 'Nominal', 'Tol -', 'Tol +', 'Min', 'Max'])
    print(df)
    # Transpozunu al
    df = df.T
    print("###############################################")
    print(df)

    # Yeni sütun ekle ve ilk satırı 'Parca No' olarak ayarla
    num_rows = len(df)
    #df.insert(0, 'Parca No', range(1, num_rows + 1))

    # Devrik dönüşüm uygula
    #df_inverse = df.apply(lambda x: 1/x if x.dtype == 'float' else x)
    

    # Yeni Excel dosyasını oluştur ve verileri kaydet
    with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a') as writer:
        df.to_excel(writer, sheet_name='TO_WORD', index=False)
     
     # Silme işlemi
        # Excel dosyasını aç
        wb = writer.book
        ws = wb['TO_WORD']
        # İlk satırı sil
        ws.delete_rows(1)

# Yeni sütun ekle ve ilk satırı 'Parca No' olarak ayarla
        ws.insert_cols(1)
        ws.cell(row=1, column=1, value='Parca No')
        # Alt satırlara doğru 'Parca No' değerlerini ekle
        for i in range(2, num_rows + 1):
            ws.cell(row=i, column=1, value=i-1)

# Kullanım örneği
input_excel = "C:/Users/alika/Desktop/Reporter_Project/Outputs/merged_data.xlsx"
output_excel = "C:/Users/alika/Desktop/Reporter_Project/Outputs/merged_data.xlsx"  # Aynı dosya üzerine kaydedilecek
apply_inverse_transformation(input_excel, output_excel)


